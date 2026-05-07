"""
Wellness Plan RAG Service - Read Excel file and provide Q&A via Claude
"""
import logging
from datetime import datetime
from typing import Optional
from anthropic import Anthropic
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class WellnessService:
    """Handles wellness plan file management and Q&A via Claude."""

    def __init__(self, api_key: str):
        """Initialize Anthropic client."""
        self.client = Anthropic(api_key=api_key)
        self.model = "claude-opus-4-7"
        self.file_cache = {}
        self.file_text = None  # Store file content as formatted text

    async def upload_file(self, file_path: str) -> str:
        """
        Load Excel file and convert to formatted text.

        Args:
            file_path: Path to the Excel file (e.g., wellness_plan_2026.xlsx)

        Returns:
            file_id: Filename (used as identifier)
        """
        try:
            file_name = file_path.split("/")[-1]
            wb = load_workbook(file_path)

            # Convert all sheets to formatted text
            text_content = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                text_content.append(f"\n{'='*60}")
                text_content.append(f"Sheet: {sheet_name}")
                text_content.append(f"{'='*60}\n")

                # Add table header and rows
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        text_content.append(" | ".join(str(cell or "") for cell in row))

            self.file_text = "\n".join(text_content)

            self.file_cache[file_name] = {
                "uploaded_at": datetime.now().isoformat(),
                "path": file_path,
                "size": len(self.file_text),
            }

            logger.info(f"✅ Loaded wellness file: {file_name}")
            return file_name

        except FileNotFoundError:
            logger.error(f"❌ File not found: {file_path}")
            raise
        except Exception as e:
            logger.error(f"❌ Error loading file: {str(e)}")
            raise

    async def answer_question(
        self, question: str, file_id: str, context: str = ""
    ) -> str:
        """
        Answer a question about the wellness plan using file content.

        Args:
            question: User's question in Hebrew or English
            file_id: ID of the wellness file (unused, uses self.file_text)
            context: Optional context for the question

        Returns:
            Answer from Claude based on the wellness plan file
        """
        try:
            if not self.file_text:
                raise ValueError("Wellness file not loaded. Call upload_file first.")

            prompt = f"""אתה עוזר תכנון הרווחה של החברה.

המשתמש שואל שאלה לגבי תוכנית הרווחה:

{question}

הנתונים של תוכנית הרווחה:
{self.file_text}

אנא בנה תשובה קצרה, מדויקת וברורה בהתבסס על הנתונים."""

            if context:
                prompt += f"\n\nהקשר נוסף: {context}"

            response = self.client.messages.create(
                model=self.model,
                max_tokens=1024,
                messages=[{"role": "user", "content": prompt}],
            )

            answer = response.content[0].text
            logger.info(f"✅ Q&A processed for: {question[:50]}...")
            return answer

        except Exception as e:
            logger.error(f"❌ Error answering question: {str(e)}")
            raise

    async def generate_new_plan(
        self, file_id: str, year: int, changes: str
    ) -> str:
        """
        Generate a new wellness plan for a given year with modifications.

        Args:
            file_id: ID of the wellness file template (unused)
            year: Target year for new plan
            changes: Requested changes (e.g., "add 50 NIS to each gift")

        Returns:
            Generated plan text with proposed changes
        """
        try:
            if not self.file_text:
                raise ValueError("Wellness file not loaded. Call upload_file first.")

            prompt = f"""אתה מעצב תוכניות רווחה.

בהתבסס על תוכנית הרווחה הקיימת, צור תוכנית חדשה לשנת {year} עם השינויים הבאים:

{changes}

הנתונים של התוכנית הקיימת:
{self.file_text}

אנא ספק:
1. סה"כ תקציב משוער
2. רשימת פעילויות מוצעות עם תאריכים
3. הערות על השינויים

תן תשובה בעברית ובפורמט ברור וקל לקריאה."""

            response = self.client.messages.create(
                model=self.model,
                max_tokens=2048,
                messages=[{"role": "user", "content": prompt}],
            )

            plan = response.content[0].text
            logger.info(f"✅ Generated new plan for {year}")
            return plan

        except Exception as e:
            logger.error(f"❌ Error generating plan: {str(e)}")
            raise

    def list_cached_files(self) -> dict:
        """List all cached wellness files."""
        return self.file_cache
